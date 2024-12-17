# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
# video_path is the path to the mouse video
# path_velocity is the path to store the velocity value for each frame
# frame_path is the path to store the velocity map for each frame
# win_size is the calculation window size for the LK sparse optical flow method,
#                            the faster the speed the larger the window selection,
#                            generally normal mice choose 35-25, IBD mice choose 15-25,
#                            depending on the specific situation.
# low_level and high_level are data filtering ranges,
#                            the low_level value is usually chosen to be 0.3-0.4,
#                            the high_level value is usually chosen to be 0.9.
# binary_value is the threshold for outlining the intestinal area of the mouse,
#                            the smaller the value, the larger the circled area,
#                            the larger the value, the smaller the circled area.
# selected is the sampling point, if selected=3 then one tracking point is selected for every three points.
# frame_rate is the number of frames per second of the video.
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

import math
import cv2
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# # All parameters that may need to be modified
video_path = r'E:\LXL\LK-flow\video\0622\control\media1.avi'
path_position = 'E:\\LXL\\LK-flow\\video\\0622\\control\\position.xlsx'
frame_path = 'E:\\LXL\\LK-flow\\video\\0622\\control\\position\\'
win_size = 45
low_level = 0.2
high_level = 0.95
binary_value = 60
selected = 2
frame_rate = 5
print(video_path)


# Excluding the top 20% and bottom 90% of data (For example min=0.2,max=0.9)
def percent_range(dataset, min, max):
    range_max = np.percentile(dataset, max * 100)
    range_min = -np.percentile(-dataset, (1 - min) * 100)

    new_data = []
    for value in dataset:
        if value < range_max and value > range_min:
            new_data.append(value)
    return new_data


# Get all pixel points inside the contours
def contours_in(contours, gray):
    p = np.zeros_like(gray)
    cv2.drawContours(p, [contours], -1, 255, -1)
    a = np.where(p == 255)[1].reshape(-1, 1)
    b = np.where(p == 255)[0].reshape(-1, 1)
    coordinate = np.concatenate([a, b], axis=1).tolist()
    return coordinate


# # Acquisition of intestinal contours from the first frame of the image
cap = cv2.VideoCapture(video_path)
ret, frame = cap.read()
gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
ret, binary = cv2.threshold(gray, binary_value, 255, cv2.THRESH_BINARY)
kernel = np.ones((1, 5), np.uint8)
binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, anchor=(2, 0), iterations=5)
contours, hierarchy = cv2.findContours(binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
contours = [cnt for cnt in contours if cv2.contourArea(cnt) > 1000 and cv2.contourArea(cnt) < 30000]
cv2.drawContours(frame, contours, -1, (0, 255, 0), 2)


# # A point in the upper right corner within the contour  was selected as the respiratory reference point
contour_number = len(contours)
X = []
Y = []
for i in range(contour_number):
    heart_contour = contours[i].reshape(-1,2)
    new_heart_contour = np.transpose(heart_contour)
    x_heart = new_heart_contour[0]
    y_heart = new_heart_contour[1]
    X.append(np.max(x_heart))
    X.append(np.min(x_heart))
    Y.append(np.max(y_heart))
    Y.append(np.min(y_heart))
heart_x_ = np.max(X)-20
heart_y_ = np.min(Y)+20
cv2.circle(frame, (heart_x_, heart_y_), 2, (90, 200, 40), thickness=-1)

cv2.imshow("img", frame)
cv2.waitKey(0)
save_path = frame_path + str(file_number) + "_0"+".png"
cv2.imwrite(save_path, frame)


# # Velocity tracking of all points within the contour using the LK sparse optical flow method
contour_number = len(contours)
prepare_list = locals()
cap = cv2.VideoCapture(video_path)
lk_params = dict(winSize=(win_size, win_size))
color = np.random.randint(0, 255, (100000, 3))
lag = round(1 / frame_rate, 3)
print(lag)
lists_dict = {}
lists_x = {}
lists_y = {}
velocity_dict = {}
all_contours = []
all_velocity = []
all_velocity_x = []
all_velocity_y = []
area = []

for i in range(contour_number):

    key = f"V{i}"
    lists_dict[key] = []
    lists_x[key] = []
    lists_y[key] = []
    contour_ = 'contour_' + str(i)
    prepare_list[contour_] = contours[i]
    contour_ = contours[i]
    contour_ = contours_in(contour_, gray)
    contour_ = np.array(contour_)
    contour_ = contour_.reshape(-1, 1, 2)

    # For larger contoured areas, one tracking point is selected for every selected points
    len_contour = len(contour_)
    area.append(int(len_contour/100))
    contour_ = contour_.reshape(len_contour, 2)
    if len_contour >= 100:
        new_contour = []
        for i, item in enumerate(contour_):
            if i % selected == 0:
                new_contour.append(item)
    else:
        new_contour = contour_

    for i, item in enumerate(new_contour):
        all_contours.append(item)
    print("追踪点个数", len(new_contour))
    new_contour.append([heart_x_, heart_y_])

    len_new_contour = len(new_contour)
    new_contour = np.array(new_contour).reshape(len_new_contour, 1, 2)   # Finalize all tracking points to be calculated

    # Speed tracing starts here
    cap = cv2.VideoCapture(video_path)
    ret, oldframe = cap.read()
    old_gray = cv2.cvtColor(oldframe, cv2.COLOR_BGR2GRAY)
    p0 = np.float32(new_contour)
    # mask = np.zeros_like(oldframe)  # Create a mask
    frame_count = 0
    while True:
        ret, frame = cap.read()
        mask = np.zeros_like(oldframe)  # A new mask is recreated for each frame, without overlaying
        frame_count += 1
        # Update tracking points for re-tracking after a certain frame rate is exceeded
        if frame_count % 10 * int(frame_rate) == 0:
            p0 = np.float32(new_contour)
        if frame is None:
            break
        # if frame_count == 360:  # Stop tracking after a certain number of frames
        #     break
        framegray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Pass in the previous frame, current frame and the tracking point of the previous frame
        p1, st, err = cv2.calcOpticalFlowPyrLK(old_gray, framegray, p0, None, **lk_params)
        p0_number = len(p0)
        scatter_v = []
        # Calculate the velocity of all tracking points except the reference point
        for m in range(p0_number-1):
            if (p1[m][0][0] - p0[m][0][0]) > 40 or (p1[m][0][1] - p0[m][0][1]) > 40:
                p1[m][0][0] = p0[m][0][0]
                p1[m][0][1] = p0[m][0][1]

            # Calculate the velocity of respiratory reference point
            v_heart_y = np.round(((p1[p0_number-1][0][1] - p0[p0_number-1][0][1]) / lag), 3)

            # Calculate the x and y direction optical flow vx,vy
            vx = np.round(((p1[m][0][0] - p0[m][0][0]) / lag), 3)
            vy = np.round(((p1[m][0][1] - p0[m][0][1]) / lag), 3)

            # Performs graded filtration of respiratory rate
            if (v_heart_y > 0 and vy > 0) or (v_heart_y < 0 and vy < 0):
                if abs(vy) > abs(v_heart_y):
                    vy = np.round(vy - v_heart_y, 3)
                elif abs(vy) < abs(v_heart_y) and abs(vy) > abs(v_heart_y) / 2:
                    vy = np.round(vy - v_heart_y * 2 / 3, 3)
                else:
                    vy = np.round(vy - v_heart_y / 2, 3)
            else:
                vy = vy

            v = np.round(math.sqrt((vx ** 2) + (vy ** 2)), 3)

            lists_x[key].append(abs(vx))
            lists_y[key].append(abs(vy))
            lists_dict[key].append(v)
            scatter_v.append(v)

        scatter_x = []
        scatter_y = []

        # # Plot the displacement trajectory of each tracking point per frame
        for i, (new, old) in enumerate(zip(p1, p0)):
            a, b = new.ravel()
            c, d = old.ravel()

            # Coordinates of the scatterplot
            scatter_x.append(int(a))
            scatter_y.append(int(b))

            if a-c > 40 or b-d > 40:
                print(i)
            mask = cv2.line(mask, (int(a), int(b)), (int(c), int(d)), color[i].tolist(), 1)  # 取整描点画线
            frame = cv2.circle(frame, (int(a), int(b)), 1, color[i].tolist(), -1)

        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        # # Draw a scatterplot of the position of each frame
        scatter_x_min = min(scatter_x)
        scatter_x_max = max(scatter_x)
        scatter_y_min = min(scatter_y)
        scatter_y_max = max(scatter_y)

        new_scatter_x = scatter_x[:-1]
        new_scatter_y = scatter_y[:-1]

        fig = plt.figure()
        ax = plt.gca()
        ax.xaxis.set_ticks_position('top')
        ax.invert_yaxis()
        plt.scatter(new_scatter_x, new_scatter_y, c=scatter_v, marker='+', cmap='jet')
        cbar = plt.colorbar()
        cbar.mappable.set_clim(0, 30)
        # cbar.set_lable('velocity',fontweight='bold')
        plt.title('Velocity scatter Plot with Color', pad=30)
        ax.set_xlabel('x-axis(pixel)', fontweight='bold')
        ax.set_ylabel('y-axis(pixel)', fontweight='bold')
        ax.set_xlim(scatter_x_min - 20, scatter_x_max + 20)
        ax.set_ylim(scatter_y_max + 20, scatter_y_min - 20)
        ax.tick_params(labelsize=10)
        every_frame = frame_path + str(frame_count) + ".png"
        plt.savefig(every_frame)
        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        # # Importing per-frame position data into a table
        workbook = load_workbook(path_position)
        sheet = workbook.active
        number = (frame_count - 1) * 4 + 1
        sheet.cell(1, number).value = frame_count
        sheet.cell(1, number + 1).value = 'x'
        for i in range(len(new_scatter_x)):
            sheet.cell(i + 2, number + 1).value = new_scatter_x[i]
        sheet.cell(1, number + 2).value = 'y'
        for i in range(len(new_scatter_y)):
            sheet.cell(i + 2, number + 2).value = new_scatter_y[i]
        sheet.cell(1, number + 3).value = 'v'
        for i in range(len(scatter_v)):
            sheet.cell(i + 2, number + 3).value = scatter_v[i]
        workbook.save(path_position)
        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

        # Stacked and updated
        img = cv2.add(frame, mask)
        cv2.imshow('frame', img)
        k = cv2.waitKey(150) & 0xff
        if k == 27:
            break
        old_gray = framegray.copy()
        p0 = p1.reshape(-1, 1, 2)

    lists_dict[key] = np.array(lists_dict[key]).reshape(-1, p0_number-1)
    velocity = np.transpose(lists_dict[key])
    lists_x[key] = np.array(lists_x[key]).reshape(-1, p0_number - 1)
    velocity_x = np.transpose(lists_x[key])
    lists_y[key] = np.array(lists_y[key]).reshape(-1, p0_number - 1)
    velocity_y = np.transpose(lists_y[key])
    velocity_dict[key] = []

    for n in range(p0_number-1):
        list_velocity = velocity[n].tolist()
        list_velocity_x = velocity_x[n].tolist()
        list_velocity_y = velocity_y[n].tolist()
        new_velocity = []
        new_velocity_x = []
        new_velocity_y = []

        # Rejecting values with excessive speed
        for i in range(len(list_velocity)):
            if list_velocity[i] < 60:
                new_velocity.append(list_velocity[i])
                new_velocity_x.append(list_velocity_x[i])
                new_velocity_y.append(list_velocity_y[i])
        array_new_velocity = np.array(new_velocity)
        array_new_velocity_x = np.array(new_velocity_x)
        array_new_velocity_y = np.array(new_velocity_y)

        # Calculate the average speed of each pixel point
        filtrate_velocity = percent_range(array_new_velocity, low_level, high_level)
        average_velocity = np.round(np.mean(filtrate_velocity), 3)
        all_velocity.append(average_velocity)
        filtrate_velocity_x = percent_range(array_new_velocity_x, low_level, high_level)
        average_velocity_x = np.round(np.mean(filtrate_velocity_x), 3)
        all_velocity_x.append(average_velocity_x)
        filtrate_velocity_y = percent_range(array_new_velocity_y, low_level, high_level)
        average_velocity_y = np.round(np.mean(filtrate_velocity_y), 3)
        all_velocity_y.append(average_velocity_y)

print()
print("Intestinal area", area)

coordinate = np.transpose(np.array(all_contours).reshape(-1, 2))
x = coordinate[0]
y = coordinate[1]
z = all_velocity

# # Calculation of individual speed indicators
high_z = []
low_z = []
max_z = max(z)
min_z = min(z)
for i,item in enumerate(z):
    if item <= max_z and item >= (0.7*max_z):
        high_z.append(item)
for i,item in enumerate(z):
    if item >= min_z and item <= (min_z/0.7):
        low_z.append(item)

average_z = np.round(np.mean(z),3)
high_v = np.round(np.mean(high_z),3)
low_v = np.round(np.mean(low_z),3)

print("Maximum velocity", max_z)
print("Minimum velocity", min_z)
print("average velocity", average_z)
print("fast zone velocity", high_v)
print("slow zone velocity", low_v)


